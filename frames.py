# frames.py

import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from datetime import datetime, timedelta
import openpyxl
from tkcalendar import DateEntry
import sqlite3

# Import functions and configurations from other modules
from database_ops import (
    get_employee_data, add_employee_to_db, update_employee_in_db, delete_employee_from_db,
    record_monthly_payroll_to_db, get_payroll_history,
    get_absences_in_month, get_overtimes_in_month,
    add_loan_to_db, get_active_loans, update_loan_remaining_amount,
    add_absence_to_db, get_absences_history,
    add_overtime_to_db, get_overtime_history,
    add_leave_to_db, get_leave_history
)
from config import ADMIN_USERNAME, ADMIN_PASSWORD, HOURLY_WORK_HOURS_IN_MONTH, OVERTIME_RATE_FACTOR, ABSENCE_RATE_FACTOR, DB_NAME


class LoginFrame(tk.Frame):
    def __init__(self, master, app_instance):
        super().__init__(master, bg="#f0f0f0") # Added a light background color
        self.app = app_instance

        # Use a sub-frame for centering content
        center_frame = tk.Frame(self, bg="#f0f0f0")
        center_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # Title
        tk.Label(center_frame, text="ورود به سیستم جامع مدیریت کارمندان", font=("Arial", 16, "bold"), bg="#f0f0f0").pack(pady=20)

        # Username
        tk.Label(center_frame, text="نام کاربری:", font=("Arial", 12), bg="#f0f0f0").pack(pady=5)
        self.username_entry = tk.Entry(center_frame, font=("Arial", 12), width=30)
        self.username_entry.pack(pady=5)
        self.username_entry.focus_set() # Set focus to username field

        # Password
        tk.Label(center_frame, text="رمز عبور:", font=("Arial", 12), bg="#f0f0f0").pack(pady=5)
        self.password_entry = tk.Entry(center_frame, show="*", font=("Arial", 12), width=30)
        self.password_entry.pack(pady=5)

        # Login Button
        tk.Button(center_frame, text="ورود", font=("Arial", 12, "bold"), command=self._perform_login, bg="#4CAF50", fg="white", padx=10, pady=5).pack(pady=20)

    def _perform_login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()

        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            messagebox.showinfo("ورود موفق", "با موفقیت وارد شدید!")
            self.app.create_main_menu_frame()
        else:
            messagebox.showerror("خطا در ورود", "نام کاربری یا رمز عبور اشتباه است.")
            self.username_entry.delete(0, tk.END)
            self.password_entry.delete(0, tk.END)
            self.username_entry.focus_set()


class MainMenuFrame(tk.Frame):
    def __init__(self, master, app_instance):
        super().__init__(master, bg="#e0f7fa")
        self.app = app_instance

        tk.Label(self, text="منوی اصلی سیستم مدیریت کارمندان", font=("Arial", 18, "bold"), bg="#e0f7fa").pack(pady=30)

        buttons_frame = tk.Frame(self, bg="#e0f7fa")
        buttons_frame.pack(pady=20)

        base_button_style = {"font": ("Arial", 14), "width": 25, "height": 2, "padx": 10, "pady": 5}

        tk.Button(buttons_frame, text="افزودن کارمند جدید", command=self.app.create_add_employee_frame, bg="#2196F3", fg="white", **base_button_style).pack(pady=10)
        tk.Button(buttons_frame, text="مشاهده لیست کارمندان", command=self.app.create_view_employees_frame, bg="#2196F3", fg="white", **base_button_style).pack(pady=10)
        tk.Button(buttons_frame, text="جستجو / ویرایش / حذف کارمند", command=self.app.create_search_edit_delete_frame, bg="#2196F3", fg="white", **base_button_style).pack(pady=10)
        tk.Button(buttons_frame, text="مدیریت حقوق و دستمزد", command=self.app.create_payroll_management_frame, bg="#8BC34A", fg="white", **base_button_style).pack(pady=10)
        tk.Button(buttons_frame, text="گزارش‌گیری (اکسل)", command=self.app.create_reports_frame, bg="#FFEB3B", fg="black", **base_button_style).pack(pady=10)
        tk.Button(buttons_frame, text="خروج از برنامه", command=self._exit_app, bg="#F44336", fg="white", **base_button_style).pack(pady=10)

        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)

    def _exit_app(self):
        if messagebox.askyesno("خروج", "آیا مطمئنید که می‌خواهید از برنامه خارج شوید؟"):
            self.master.quit()


class AddEmployeeFrame(tk.Frame):
    def __init__(self, master, app_instance):
        super().__init__(master, bg="#f9f9f9")
        self.app = app_instance

        tk.Label(self, text="افزودن کارمند جدید", font=("Arial", 16, "bold"), bg="#f9f9f9").pack(pady=20)

        self.entries = {}
        fields = [
            ("نام:", "first_name"),
            ("نام خانوادگی:", "last_name"),
            ("کد ملی (یکتا):", "national_id"),
            ("سمت:", "position"),
            ("حقوق ثابت:", "base_salary")
        ]

        input_frame = tk.Frame(self, bg="#f9f9f9")
        input_frame.pack(pady=10, padx=50)

        for i, (text, key) in enumerate(fields):
            tk.Label(input_frame, text=text, font=("Arial", 12), bg="#f9f9f9").grid(row=i, column=1, padx=10, pady=5, sticky="w")
            entry = tk.Entry(input_frame, font=("Arial", 12), width=30)
            entry.grid(row=i, column=0, padx=10, pady=5, sticky="ew")
            self.entries[key] = entry

        input_frame.grid_columnconfigure(0, weight=1)

        buttons_frame = tk.Frame(self, bg="#f9f9f9")
        buttons_frame.pack(pady=20)

        tk.Button(buttons_frame, text="ذخیره کارمند", font=("Arial", 12, "bold"), command=self._save_employee, bg="#4CAF50", fg="white", padx=10, pady=5).pack(side=tk.LEFT, padx=10)
        tk.Button(buttons_frame, text="بازگشت به منو", font=("Arial", 12), command=self.app.create_main_menu_frame, bg="#FFC107", fg="black", padx=10, pady=5).pack(side=tk.RIGHT, padx=10)


    def _save_employee(self):
        data = {key: entry.get().strip() for key, entry in self.entries.items()}

        for key, value in data.items():
            if not value:
                messagebox.showwarning("ورودی ناقص", "لطفا تمام فیلدها را پر کنید.")
                return

        national_id = data["national_id"]
        if not national_id.isdigit() or len(national_id) != 10:
            messagebox.showerror("خطا", "کد ملی باید یک عدد ۱۰ رقمی باشد.")
            return

        try:
            base_salary = float(data["base_salary"])
            if base_salary < 0:
                 messagebox.showerror("خطا", "حقوق ثابت نمی‌تواند منفی باشد.")
                 return
        except ValueError:
            messagebox.showerror("خطا", "حقوق ثابت باید یک عدد معتبر باشد.")
            return

        success = add_employee_to_db(national_id, data["first_name"], data["last_name"], data["position"], base_salary)
        if success:
            messagebox.showinfo("موفقیت", f"کارمند {data['first_name']} {data['last_name']} با موفقیت اضافه شد.")
            for entry in self.entries.values():
                entry.delete(0, tk.END)
        else:
            messagebox.showerror("خطا", "کارمندی با این کد ملی از قبل وجود دارد.")

class ViewEmployeesFrame(tk.Frame):
    def __init__(self, master, app_instance):
        super().__init__(master, bg="#f9f9f9")
        self.app = app_instance

        tk.Label(self, text="لیست کارمندان", font=("Arial", 16, "bold"), bg="#f9f9f9").pack(pady=20)

        self.tree = ttk.Treeview(self, columns=("national_id", "first_name", "last_name", "position", "base_salary"), show="headings")
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.tree.heading("national_id", text="کد ملی", anchor="center")
        self.tree.heading("first_name", text="نام", anchor="center")
        self.tree.heading("last_name", text="نام خانوادگی", anchor="center")
        self.tree.heading("position", text="سمت", anchor="center")
        self.tree.heading("base_salary", text="حقوق ثابت", anchor="center")

        self.tree.column("national_id", width=100, anchor="center")
        self.tree.column("first_name", width=100, anchor="center")
        self.tree.column("last_name", width=120, anchor="center")
        self.tree.column("position", width=120, anchor="center")
        self.tree.column("base_salary", width=100, anchor="center")

        self._load_employees_to_tree()

        yscrollbar = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        yscrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=yscrollbar.set)

        tk.Button(self, text="بازگشت به منو", font=("Arial", 12), command=self.app.create_main_menu_frame, bg="#FFC107", fg="black", padx=10, pady=5).pack(pady=10)

    def _load_employees_to_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        employees_data = get_employee_data()
        if not employees_data:
            self.tree.insert("", tk.END, values=("", "", "هیچ کارمندی در سیستم ثبت نشده است.", "", ""), tags=('no_data',))
            self.tree.tag_configure('no_data', foreground='gray', font=('Arial', 10, 'italic'))
            return

        for emp_row in employees_data:
            self.tree.insert("", tk.END, values=(
                emp_row[0], # national_id
                emp_row[1], # first_name
                emp_row[2], # last_name
                emp_row[3], # position
                f"{emp_row[4]:.0f}" # base_salary
            ))

class SearchEditDeleteFrame(tk.Frame):
    def __init__(self, master, app_instance):
        super().__init__(master, bg="#f9f9f9")
        self.app = app_instance
        self.current_employee_id = None

        tk.Label(self, text="جستجو، ویرایش و حذف کارمند", font=("Arial", 16, "bold"), bg="#f9f9f9").pack(pady=20)

        # Search Frame
        search_frame = tk.LabelFrame(self, text="جستجو", bg="#f9f9f9", padx=10, pady=10)
        search_frame.pack(pady=10, padx=20, fill=tk.X)

        tk.Label(search_frame, text="کد ملی / نام / نام خانوادگی:", font=("Arial", 12), bg="#f9f9f9").pack(side=tk.RIGHT, padx=5)
        self.search_entry = tk.Entry(search_frame, font=("Arial", 12), width=30)
        self.search_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X, padx=5)
        tk.Button(search_frame, text="جستجو", font=("Arial", 12), command=self._search_employee, bg="#03A9F4", fg="white", padx=10, pady=5).pack(side=tk.LEFT, padx=5)

        # Search Results Treeview
        self.search_results_tree = ttk.Treeview(self, columns=("national_id", "first_name", "last_name", "position"), show="headings")
        self.search_results_tree.pack(fill=tk.X, padx=20, pady=5)
        self.search_results_tree.heading("national_id", text="کد ملی", anchor="center")
        self.search_results_tree.heading("first_name", text="نام", anchor="center")
        self.search_results_tree.heading("last_name", text="نام خانوادگی", anchor="center")
        self.search_results_tree.heading("position", text="سمت", anchor="center")
        self.search_results_tree.column("national_id", width=100, anchor="center")
        self.search_results_tree.column("first_name", width=100, anchor="center")
        self.search_results_tree.column("last_name", width=120, anchor="center")
        self.search_results_tree.column("position", width=120, anchor="center")
        self.search_results_tree.bind("<<TreeviewSelect>>", self._on_employee_select) # Bind selection event

        # Details Frame
        details_frame = tk.LabelFrame(self, text="اطلاعات کارمند", bg="#f9f9f9", padx=10, pady=10)
        details_frame.pack(pady=10, padx=20, fill=tk.X)

        self.edit_entries = {}
        fields = [
            ("نام:", "first_name"),
            ("نام خانوادگی:", "last_name"),
            ("سمت:", "position"),
            ("حقوق ثابت:", "base_salary")
        ]

        for i, (text, key) in enumerate(fields):
            tk.Label(details_frame, text=text, font=("Arial", 10), bg="#f9f9f9").grid(row=i, column=1, padx=5, pady=2, sticky="w")
            entry = tk.Entry(details_frame, font=("Arial", 10), width=30)
            entry.grid(row=i, column=0, padx=5, pady=2, sticky="ew")
            self.edit_entries[key] = entry

        details_frame.grid_columnconfigure(0, weight=1)

        action_buttons_frame = tk.Frame(self, bg="#f9f9f9")
        action_buttons_frame.pack(pady=10)

        tk.Button(action_buttons_frame, text="ویرایش", font=("Arial", 12, "bold"), command=self._edit_employee, bg="#FF9800", fg="white", padx=10, pady=5).pack(side=tk.LEFT, padx=10)
        tk.Button(action_buttons_frame, text="حذف", font=("Arial", 12, "bold"), command=self._delete_employee, bg="#F44336", fg="white", padx=10, pady=5).pack(side=tk.LEFT, padx=10)
        tk.Button(action_buttons_frame, text="بازگشت به منو", font=("Arial", 12), command=self.app.create_main_menu_frame, bg="#FFC107", fg="black", padx=10, pady=5).pack(side=tk.RIGHT, padx=10)

        self._clear_edit_fields()

    def _search_employee(self):
        search_query = self.search_entry.get().strip()
        if not search_query:
            messagebox.showwarning("ورودی ناقص", "لطفا کد ملی، نام یا نام خانوادگی را برای جستجو وارد کنید.")
            self._clear_edit_fields()
            self._clear_search_results()
            return

        self._clear_search_results()
        self._clear_edit_fields()

        # Using direct sqlite3 connection here because database_ops functions only fetch by national_id directly.
        # Could add a search function to database_ops if needed frequently.
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()

        # Search by national_id, first_name, or last_name
        cursor.execute("SELECT national_id, first_name, last_name, position FROM employees WHERE national_id LIKE ? OR first_name LIKE ? OR last_name LIKE ?",
                       (f"%{search_query}%", f"%{search_query}%", f"%{search_query}%"))
        results = cursor.fetchall()
        conn.close()

        if results:
            for emp_row in results:
                self.search_results_tree.insert("", tk.END, values=emp_row)
            if len(results) == 1:
                # If only one result, auto-select it and load details
                self.search_results_tree.selection_set(self.search_results_tree.get_children()[0])
                self._on_employee_select(None) # Call the select handler
        else:
            messagebox.showinfo("یافت نشد", "کارمندی با این مشخصات یافت نشد.")


    def _on_employee_select(self, event):
        selected_item = self.search_results_tree.selection()
        if not selected_item:
            self.current_employee_id = None
            self._clear_edit_fields()
            return

        # Get data from the selected row
        item_values = self.search_results_tree.item(selected_item[0], 'values')
        national_id = item_values[0] # national_id is the first value

        emp_data = get_employee_data(national_id)
        if emp_data:
            self.current_employee_id = national_id
            self._fill_edit_fields(emp_data)
        else:
            # This case should ideally not happen if data is consistent
            messagebox.showerror("خطا", "مشکلی در بارگذاری اطلاعات کارمند رخ داد.")
            self.current_employee_id = None
            self._clear_edit_fields()


    def _fill_edit_fields(self, emp_data):
        self.edit_entries["first_name"].delete(0, tk.END)
        self.edit_entries["first_name"].insert(0, emp_data[1]) # first_name
        self.edit_entries["last_name"].delete(0, tk.END)
        self.edit_entries["last_name"].insert(0, emp_data[2])  # last_name
        self.edit_entries["position"].delete(0, tk.END)
        self.edit_entries["position"].insert(0, emp_data[3])  # position
        self.edit_entries["base_salary"].delete(0, tk.END)
        self.edit_entries["base_salary"].insert(0, str(emp_data[4])) # base_salary

    def _clear_edit_fields(self):
        for entry in self.edit_entries.values():
            entry.delete(0, tk.END)
        self.current_employee_id = None

    def _clear_search_results(self):
        for item in self.search_results_tree.get_children():
            self.search_results_tree.delete(item)

    def _edit_employee(self):
        if not self.current_employee_id:
            messagebox.showwarning("هشدار", "لطفا ابتدا یک کارمند را جستجو و از لیست انتخاب کنید تا اطلاعات آن را ویرایش کنید.")
            return

        data = {key: entry.get().strip() for key, entry in self.edit_entries.items()}

        for key, value in data.items():
            if not value:
                messagebox.showwarning("ورودی ناقص", "لطفا تمام فیلدها را پر کنید.")
                return

        try:
            base_salary = float(data["base_salary"])
            if base_salary < 0:
                 messagebox.showerror("خطا", "حقوق ثابت نمی‌تواند منفی باشد.")
                 return
        except ValueError:
            messagebox.showerror("خطا", "حقوق ثابت باید یک عدد معتبر باشد.")
            return

        update_employee_in_db(self.current_employee_id, data["first_name"], data["last_name"], data["position"], base_salary)
        messagebox.showinfo("موفقیت", f"اطلاعات کارمند با کد ملی {self.current_employee_id} با موفقیت به‌روزرسانی شد.")
        self._clear_edit_fields()
        self._clear_search_results()
        self.search_entry.delete(0, tk.END)


    def _delete_employee(self):
        if not self.current_employee_id:
            messagebox.showwarning("هشدار", "لطفا ابتدا یک کارمند را جستجو و از لیست انتخاب کنید تا آن را حذف کنید.")
            return

        if messagebox.askyesno("تایید حذف", f"آیا مطمئنید که می‌خواهید کارمند با کد ملی {self.current_employee_id} و تمامی سوابق مربوطه (حقوق، غیبت، اضافه کار، مرخصی، وام) را حذف کنید؟ این عمل برگشت‌ناپذیر است."):
            delete_employee_from_db(self.current_employee_id)
            messagebox.showinfo("موفقیت", "کارمند با موفقیت حذف شد.")
            self._clear_edit_fields()
            self._clear_search_results()
            self.search_entry.delete(0, tk.END)


class AbsenceOvertimeLeaveFrame(tk.Frame):
    def __init__(self, master, app_instance, employee_national_id):
        super().__init__(master, bg="#f9f9f9")
        self.app = app_instance
        self.employee_national_id = employee_national_id

        # Fetch employee info for display
        employee_data = get_employee_data(self.employee_national_id)
        if employee_data:
            self.employee_name = f"{employee_data[1]} {employee_data[2]} (کد ملی: {employee_data[0]})"
        else:
            self.employee_name = "کارمند نامشخص" # Should not happen if called correctly

        # Top Bar for Title and Back Button
        top_bar_frame = tk.Frame(self, bg="#e0f7fa")
        top_bar_frame.pack(fill=tk.X, padx=10, pady=5)
        tk.Button(top_bar_frame, text="بازگشت به مدیریت حقوق", font=("Arial", 11), command=self.app.create_payroll_management_frame, bg="#FFC107", fg="black", padx=8, pady=3).pack(side=tk.LEFT, padx=5)
        tk.Label(top_bar_frame, text=f"مدیریت غیبت، اضافه کار و مرخصی برای: {self.employee_name}", font=("Arial", 14, "bold"), bg="#e0f7fa").pack(side=tk.RIGHT, padx=10)

        # Notebook (Tabbed Interface)
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

        # --- Absence Tab ---
        self.absence_frame = tk.Frame(self.notebook, bg="#f9f9f9")
        self.notebook.add(self.absence_frame, text="مدیریت غیبت")
        self._setup_absence_tab(self.absence_frame)

        # --- Overtime Tab ---
        self.overtime_frame = tk.Frame(self.notebook, bg="#f9f9f9")
        self.notebook.add(self.overtime_frame, text="مدیریت اضافه کار")
        self._setup_overtime_tab(self.overtime_frame)

        # --- Leave Tab ---
        self.leave_frame = tk.Frame(self.notebook, bg="#f9f9f9")
        self.notebook.add(self.leave_frame, text="مدیریت مرخصی")
        self._setup_leave_tab(self.leave_frame)

        # --- Loans Tab (NEW) ---
        self.loans_frame = tk.Frame(self.notebook, bg="#f9f9f9")
        self.notebook.add(self.loans_frame, text="مدیریت وام / پیش‌پرداخت")
        self._setup_loans_tab(self.loans_frame)


        # Initial load of data for the first tab
        self._load_absences_history()
        # Bind for re-loading data on tab change
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_change)


    def _setup_absence_tab(self, parent_frame):
        # Input Section
        input_frame = tk.LabelFrame(parent_frame, text="ثبت غیبت جدید", bg="#f9f9f9", padx=10, pady=10)
        input_frame.pack(pady=10, padx=20, fill=tk.X)

        tk.Label(input_frame, text="تاریخ غیبت:", font=("Arial", 10), bg="#f9f9f9").grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.abs_date_entry = DateEntry(input_frame, selectmode='day', date_pattern='yyyy-mm-dd', font=("Arial", 10), width=15)
        self.abs_date_entry.grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        tk.Label(input_frame, text="مدت غیبت (ساعت):", font=("Arial", 10), bg="#f9f9f9").grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.abs_hours_entry = tk.Entry(input_frame, font=("Arial", 10), width=10)
        self.abs_hours_entry.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        tk.Label(input_frame, text="دلیل غیبت:", font=("Arial", 10), bg="#f9f9f9").grid(row=1, column=3, padx=5, pady=5, sticky="w")
        self.abs_reason_entry = tk.Entry(input_frame, font=("Arial", 10), width=30)
        self.abs_reason_entry.grid(row=1, column=0, columnspan=3, padx=5, pady=5, sticky="ew")

        tk.Button(input_frame, text="ثبت غیبت", font=("Arial", 12, "bold"), command=self._record_absence, bg="#FF5722", fg="white", padx=10, pady=5).grid(row=2, column=0, columnspan=4, pady=10)
        input_frame.grid_columnconfigure(0, weight=1)

        # History Section
        history_frame = tk.LabelFrame(parent_frame, text="سوابق غیبت", bg="#f9f9f9", padx=10, pady=10)
        history_frame.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)

        self.absence_history_tree = ttk.Treeview(history_frame, columns=("date", "hours", "reason"), show="headings")
        self.absence_history_tree.pack(fill=tk.BOTH, expand=True)

        self.absence_history_tree.heading("date", text="تاریخ غیبت", anchor="center")
        self.absence_history_tree.heading("hours", text="مدت (ساعت)", anchor="center")
        self.absence_history_tree.heading("reason", text="دلیل", anchor="center")

        self.absence_history_tree.column("date", width=100, anchor="center")
        self.absence_history_tree.column("hours", width=80, anchor="center")
        self.absence_history_tree.column("reason", width=250, anchor="w")

        history_yscrollbar = ttk.Scrollbar(history_frame, orient="vertical", command=self.absence_history_tree.yview)
        history_yscrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.absence_history_tree.configure(yscrollcommand=history_yscrollbar.set)


    def _record_absence(self):
        absence_date_str = self.abs_date_entry.get() # Get from DateEntry
        hours_absent_str = self.abs_hours_entry.get().strip()
        reason = self.abs_reason_entry.get().strip()

        if not all([absence_date_str, hours_absent_str]):
            messagebox.showwarning("ورودی ناقص", "لطفا تاریخ و مدت غیبت را وارد کنید.")
            return

        try:
            # Validate date format (DateEntry usually handles this, but good for safety)
            datetime.strptime(absence_date_str, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("خطا", "فرمت تاریخ نامعتبر است.")
            return

        try:
            hours_absent = float(hours_absent_str)
            if hours_absent <= 0:
                messagebox.showerror("خطا", "مدت غیبت باید عددی مثبت باشد.")
                return
        except ValueError:
            messagebox.showerror("خطا", "مدت غیبت باید یک عدد معتبر باشد.")
            return

        add_absence_to_db(self.employee_national_id, absence_date_str, hours_absent, reason)
        messagebox.showinfo("موفقیت", "غیبت با موفقیت ثبت شد.")
        self.abs_hours_entry.delete(0, tk.END)
        self.abs_reason_entry.delete(0, tk.END)
        # self.abs_date_entry.set_date(datetime.now()) # Reset DateEntry

        self._load_absences_history()

    def _load_absences_history(self):
        for item in self.absence_history_tree.get_children():
            self.absence_history_tree.delete(item)

        history = get_absences_history(self.employee_national_id)
        if not history:
            self.absence_history_tree.insert("", tk.END, values=("", "هیچ سابقه غیبت یافت نشد.", ""), tags=('no_data',))
            self.absence_history_tree.tag_configure('no_data', foreground='gray', font=('Arial', 10, 'italic'))
            return

        for record in history:
            self.absence_history_tree.insert("", tk.END, values=(record[0], f"{record[1]:.1f}", record[2]))

    def _setup_overtime_tab(self, parent_frame):
        # Input Section
        input_frame = tk.LabelFrame(parent_frame, text="ثبت اضافه کار جدید", bg="#f9f9f9", padx=10, pady=10)
        input_frame.pack(pady=10, padx=20, fill=tk.X)

        tk.Label(input_frame, text="تاریخ اضافه کار:", font=("Arial", 10), bg="#f9f9f9").grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.ot_date_entry = DateEntry(input_frame, selectmode='day', date_pattern='yyyy-mm-dd', font=("Arial", 10), width=15)
        self.ot_date_entry.grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        tk.Label(input_frame, text="مدت اضافه کار (ساعت):", font=("Arial", 10), bg="#f9f9f9").grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.ot_hours_entry = tk.Entry(input_frame, font=("Arial", 10), width=10)
        self.ot_hours_entry.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        tk.Label(input_frame, text="توضیحات:", font=("Arial", 10), bg="#f9f9f9").grid(row=1, column=3, padx=5, pady=5, sticky="w")
        self.ot_description_entry = tk.Entry(input_frame, font=("Arial", 10), width=30)
        self.ot_description_entry.grid(row=1, column=0, columnspan=3, padx=5, pady=5, sticky="ew")

        tk.Button(input_frame, text="ثبت اضافه کار", font=("Arial", 12, "bold"), command=self._record_overtime, bg="#4CAF50", fg="white", padx=10, pady=5).grid(row=2, column=0, columnspan=4, pady=10)
        input_frame.grid_columnconfigure(0, weight=1)

        # History Section
        history_frame = tk.LabelFrame(parent_frame, text="سوابق اضافه کار", bg="#f9f9f9", padx=10, pady=10)
        history_frame.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)

        self.overtime_history_tree = ttk.Treeview(history_frame, columns=("date", "hours", "description"), show="headings")
        self.overtime_history_tree.pack(fill=tk.BOTH, expand=True)

        self.overtime_history_tree.heading("date", text="تاریخ اضافه کار", anchor="center")
        self.overtime_history_tree.heading("hours", text="مدت (ساعت)", anchor="center")
        self.overtime_history_tree.heading("description", text="توضیحات", anchor="center")

        self.overtime_history_tree.column("date", width=100, anchor="center")
        self.overtime_history_tree.column("hours", width=80, anchor="center")
        self.overtime_history_tree.column("description", width=250, anchor="w")

        history_yscrollbar = ttk.Scrollbar(history_frame, orient="vertical", command=self.overtime_history_tree.yview)
        history_yscrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.overtime_history_tree.configure(yscrollcommand=history_yscrollbar.set)

    def _record_overtime(self):
        overtime_date_str = self.ot_date_entry.get() # Get from DateEntry
        hours_worked_str = self.ot_hours_entry.get().strip()
        description = self.ot_description_entry.get().strip()

        if not all([overtime_date_str, hours_worked_str]):
            messagebox.showwarning("ورودی ناقص", "لطفا تاریخ و مدت اضافه کار را وارد کنید.")
            return

        try:
            datetime.strptime(overtime_date_str, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("خطا", "فرمت تاریخ نامعتبر است.")
            return

        try:
            hours_worked = float(hours_worked_str)
            if hours_worked <= 0:
                messagebox.showerror("خطا", "مدت اضافه کار باید عددی مثبت باشد.")
                return
        except ValueError:
            messagebox.showerror("خطا", "مدت اضافه کار باید یک عدد معتبر باشد.")
            return

        add_overtime_to_db(self.employee_national_id, overtime_date_str, hours_worked, description)
        messagebox.showinfo("موفقیت", "اضافه کار با موفقیت ثبت شد.")
        self.ot_hours_entry.delete(0, tk.END)
        self.ot_description_entry.delete(0, tk.END)
        # self.ot_date_entry.set_date(datetime.now()) # Reset DateEntry

        self._load_overtime_history()

    def _load_overtime_history(self):
        for item in self.overtime_history_tree.get_children():
            self.overtime_history_tree.delete(item)

        history = get_overtime_history(self.employee_national_id)
        if not history:
            self.overtime_history_tree.insert("", tk.END, values=("", "هیچ سابقه اضافه کار یافت نشد.", ""), tags=('no_data',))
            self.overtime_history_tree.tag_configure('no_data', foreground='gray', font=('Arial', 10, 'italic'))
            return

        for record in history:
            self.overtime_history_tree.insert("", tk.END, values=(record[0], f"{record[1]:.1f}", record[2]))

    def _setup_leave_tab(self, parent_frame):
        # Input Section
        input_frame = tk.LabelFrame(parent_frame, text="ثبت مرخصی جدید", bg="#f9f9f9", padx=10, pady=10)
        input_frame.pack(pady=10, padx=20, fill=tk.X)

        tk.Label(input_frame, text="تاریخ شروع:", font=("Arial", 10), bg="#f9f9f9").grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.leave_start_date_entry = DateEntry(input_frame, selectmode='day', date_pattern='yyyy-mm-dd', font=("Arial", 10), width=15)
        self.leave_start_date_entry.grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        tk.Label(input_frame, text="تاریخ پایان:", font=("Arial", 10), bg="#f9f9f9").grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.leave_end_date_entry = DateEntry(input_frame, selectmode='day', date_pattern='yyyy-mm-dd', font=("Arial", 10), width=15)
        self.leave_end_date_entry.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        tk.Label(input_frame, text="نوع مرخصی:", font=("Arial", 10), bg="#f9f9f9").grid(row=1, column=3, padx=5, pady=5, sticky="w")
        self.leave_type_combobox = ttk.Combobox(input_frame, values=["استحقاقی", "استعلاجی", "بدون حقوق", "ساعتی"], font=("Arial", 10), width=15, state="readonly")
        self.leave_type_combobox.set("استحقاقی") # Default value
        self.leave_type_combobox.grid(row=1, column=2, padx=5, pady=5, sticky="ew")

        tk.Label(input_frame, text="توضیحات:", font=("Arial", 10), bg="#f9f9f9").grid(row=2, column=3, padx=5, pady=5, sticky="w")
        self.leave_description_entry = tk.Entry(input_frame, font=("Arial", 10), width=30)
        self.leave_description_entry.grid(row=2, column=0, columnspan=3, padx=5, pady=5, sticky="ew")

        tk.Button(input_frame, text="ثبت مرخصی", font=("Arial", 12, "bold"), command=self._record_leave, bg="#673AB7", fg="white", padx=10, pady=5).grid(row=3, column=0, columnspan=4, pady=10)
        input_frame.grid_columnconfigure(0, weight=1)


        # History Section
        history_frame = tk.LabelFrame(parent_frame, text="سوابق مرخصی", bg="#f9f9f9", padx=10, pady=10)
        history_frame.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)

        self.leave_history_tree = ttk.Treeview(history_frame, columns=("start_date", "end_date", "type", "duration", "description"), show="headings")
        self.leave_history_tree.pack(fill=tk.BOTH, expand=True)

        self.leave_history_tree.heading("start_date", text="تاریخ شروع", anchor="center")
        self.leave_history_tree.heading("end_date", text="تاریخ پایان", anchor="center")
        self.leave_history_tree.heading("type", text="نوع مرخصی", anchor="center")
        self.leave_history_tree.heading("duration", text="مدت (روز)", anchor="center")
        self.leave_history_tree.heading("description", text="توضیحات", anchor="center")

        self.leave_history_tree.column("start_date", width=100, anchor="center")
        self.leave_history_tree.column("end_date", width=100, anchor="center")
        self.leave_history_tree.column("type", width=80, anchor="center")
        self.leave_history_tree.column("duration", width=60, anchor="center")
        self.leave_history_tree.column("description", width=200, anchor="w")

        history_yscrollbar = ttk.Scrollbar(history_frame, orient="vertical", command=self.leave_history_tree.yview)
        history_yscrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.leave_history_tree.configure(yscrollcommand=history_yscrollbar.set)

    def _record_leave(self):
        start_date_str = self.leave_start_date_entry.get() # From DateEntry
        end_date_str = self.leave_end_date_entry.get()     # From DateEntry
        leave_type = self.leave_type_combobox.get().strip()
        description = self.leave_description_entry.get().strip()

        if not all([start_date_str, end_date_str, leave_type]):
            messagebox.showwarning("ورودی ناقص", "لطفا تمام فیلدهای تاریخ شروع، پایان و نوع مرخصی را پر کنید.")
            return

        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("خطا", "فرمت تاریخ نامعتبر است.")
            return

        if start_date > end_date:
            messagebox.showerror("خطا", "تاریخ شروع مرخصی نمی‌تواند بعد از تاریخ پایان باشد.")
            return

        duration_days = (end_date - start_date).days + 1 # +1 to include start day

        add_leave_to_db(self.employee_national_id, start_date_str, end_date_str, leave_type, duration_days, description)
        messagebox.showinfo("موفقیت", "مرخصی با موفقیت ثبت شد.")
        self.leave_description_entry.delete(0, tk.END)
        # self.leave_start_date_entry.set_date(datetime.now()) # Reset DateEntry
        # self.leave_end_date_entry.set_date(datetime.now())   # Reset DateEntry

        self._load_leave_history()


    def _load_leave_history(self):
        for item in self.leave_history_tree.get_children():
            self.leave_history_tree.delete(item)

        history = get_leave_history(self.employee_national_id)
        if not history:
            self.leave_history_tree.insert("", tk.END, values=("", "", "هیچ سابقه مرخصی یافت نشد.", "", ""), tags=('no_data',))
            self.leave_history_tree.tag_configure('no_data', foreground='gray', font=('Arial', 10, 'italic'))
            return

        for record in history:
            self.leave_history_tree.insert("", tk.END, values=(record[0], record[1], record[2], f"{record[3]:.1f}", record[4]))

    def _setup_loans_tab(self, parent_frame):
        # Input Section
        input_frame = tk.LabelFrame(parent_frame, text="ثبت وام / پیش‌پرداخت جدید", bg="#f9f9f9", padx=10, pady=10)
        input_frame.pack(pady=10, padx=20, fill=tk.X)

        tk.Label(input_frame, text="تاریخ وام:", font=("Arial", 10), bg="#f9f9f9").grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.loan_date_entry = DateEntry(input_frame, selectmode='day', date_pattern='yyyy-mm-dd', font=("Arial", 10), width=15)
        self.loan_date_entry.grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        tk.Label(input_frame, text="مبلغ وام:", font=("Arial", 10), bg="#f9f9f9").grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.loan_amount_entry = tk.Entry(input_frame, font=("Arial", 10), width=15)
        self.loan_amount_entry.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        tk.Label(input_frame, text="مبلغ قسط پیشنهادی (ماهانه):", font=("Arial", 10), bg="#f9f9f9").grid(row=1, column=3, padx=5, pady=5, sticky="w")
        self.loan_installment_entry = tk.Entry(input_frame, font=("Arial", 10), width=15)
        self.loan_installment_entry.grid(row=1, column=2, padx=5, pady=5, sticky="ew")

        tk.Label(input_frame, text="توضیحات:", font=("Arial", 10), bg="#f9f9f9").grid(row=2, column=3, padx=5, pady=5, sticky="w")
        self.loan_description_entry = tk.Entry(input_frame, font=("Arial", 10), width=30)
        self.loan_description_entry.grid(row=2, column=0, columnspan=3, padx=5, pady=5, sticky="ew")

        tk.Button(input_frame, text="ثبت وام", font=("Arial", 12, "bold"), command=self._record_loan, bg="#FF9800", fg="white", padx=10, pady=5).grid(row=3, column=0, columnspan=4, pady=10)
        input_frame.grid_columnconfigure(0, weight=1)

        # History Section
        history_frame = tk.LabelFrame(parent_frame, text="وام‌ها / پیش‌پرداخت‌های فعال", bg="#f9f9f9", padx=10, pady=10)
        history_frame.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)

        self.loans_history_tree = ttk.Treeview(history_frame, columns=("id", "date", "amount", "remaining", "installment", "description"), show="headings")
        self.loans_history_tree.pack(fill=tk.BOTH, expand=True)

        self.loans_history_tree.heading("id", text="شناسه", anchor="center")
        self.loans_history_tree.heading("date", text="تاریخ وام", anchor="center")
        self.loans_history_tree.heading("amount", text="مبلغ کل", anchor="center")
        self.loans_history_tree.heading("remaining", text="باقیمانده", anchor="center")
        self.loans_history_tree.heading("installment", text="قسط ماهانه", anchor="center")
        self.loans_history_tree.heading("description", text="توضیحات", anchor="center")

        self.loans_history_tree.column("id", width=50, anchor="center")
        self.loans_history_tree.column("date", width=100, anchor="center")
        self.loans_history_tree.column("amount", width=100, anchor="center")
        self.loans_history_tree.column("remaining", width=100, anchor="center")
        self.loans_history_tree.column("installment", width=100, anchor="center")
        self.loans_history_tree.column("description", width=200, anchor="w")

        history_yscrollbar = ttk.Scrollbar(history_frame, orient="vertical", command=self.loans_history_tree.yview)
        history_yscrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.loans_history_tree.configure(yscrollcommand=history_yscrollbar.set)

    def _record_loan(self):
        loan_date_str = self.loan_date_entry.get()
        amount_str = self.loan_amount_entry.get().strip()
        installment_str = self.loan_installment_entry.get().strip()
        description = self.loan_description_entry.get().strip()

        if not all([loan_date_str, amount_str]):
            messagebox.showwarning("ورودی ناقص", "لطفا تاریخ و مبلغ وام را وارد کنید.")
            return

        try:
            datetime.strptime(loan_date_str, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("خطا", "فرمت تاریخ نامعتبر است.")
            return

        try:
            amount = float(amount_str)
            if amount <= 0:
                messagebox.showerror("خطا", "مبلغ وام باید عددی مثبت باشد.")
                return
        except ValueError:
            messagebox.showerror("خطا", "مبلغ وام باید یک عدد معتبر باشد.")
            return

        installment_amount = 0
        if installment_str:
            try:
                installment_amount = float(installment_str)
                if installment_amount < 0:
                    messagebox.showerror("خطا", "مبلغ قسط نمی‌تواند منفی باشد.")
                    return
            except ValueError:
                messagebox.showerror("خطا", "مبلغ قسط باید یک عدد معتبر باشد.")
                return

        add_loan_to_db(self.employee_national_id, loan_date_str, amount, installment_amount, description)
        messagebox.showinfo("موفقیت", "وام / پیش‌پرداخت با موفقیت ثبت شد.")
        self.loan_amount_entry.delete(0, tk.END)
        self.loan_installment_entry.delete(0, tk.END)
        self.loan_description_entry.delete(0, tk.END)
        self._load_loans_history()


    def _load_loans_history(self):
        for item in self.loans_history_tree.get_children():
            self.loans_history_tree.delete(item)

        history = get_active_loans(self.employee_national_id)
        if not history:
            self.loans_history_tree.insert("", tk.END, values=("", "هیچ وام فعال یافت نشد.", "", "", "", ""), tags=('no_data',))
            self.loans_history_tree.tag_configure('no_data', foreground='gray', font=('Arial', 10, 'italic'))
            return

        for record in history:
            self.loans_history_tree.insert("", tk.END, values=(record[0], record[1], f"{record[2]:.0f}", f"{record[3]:.0f}", f"{record[4]:.0f}", record[5]))


    def _on_tab_change(self, event):
        """Called when a tab is selected to refresh its data."""
        selected_tab = self.notebook.tab(self.notebook.select(), "text")
        if selected_tab == "مدیریت غیبت":
            self._load_absences_history()
        elif selected_tab == "مدیریت اضافه کار":
            self._load_overtime_history()
        elif selected_tab == "مدیریت مرخصی":
            self._load_leave_history()
        elif selected_tab == "مدیریت وام / پیش‌پرداخت":
            self._load_loans_history()


class PayrollManagementFrame(tk.Frame):
    def __init__(self, master, app_instance):
        super().__init__(master, bg="#f9f9f9")
        self.app = app_instance
        self.current_employee_id = None

        # Create a frame for the title and back button at the top
        top_bar_frame = tk.Frame(self, bg="#e0f7fa")
        top_bar_frame.pack(fill=tk.X, padx=10, pady=5)

        # Back button at the top-left
        tk.Button(top_bar_frame, text="بازگشت به منو", font=("Arial", 11), command=self.app.create_main_menu_frame, bg="#FFC107", fg="black", padx=8, pady=3).pack(side=tk.LEFT, padx=5)
        # Title can be placed to the right of the button
        tk.Label(top_bar_frame, text="مدیریت حقوق و دستمزد", font=("Arial", 16, "bold"), bg="#e0f7fa").pack(side=tk.RIGHT, padx=10)


        # Search Employee Section
        search_frame = tk.LabelFrame(self, text="جستجوی کارمند", bg="#f9f9f9", padx=10, pady=10)
        search_frame.pack(pady=10, padx=20, fill=tk.X)
        tk.Label(search_frame, text="کد ملی کارمند:", font=("Arial", 12), bg="#f9f9f9").pack(side=tk.RIGHT, padx=5)
        self.search_entry = tk.Entry(search_frame, font=("Arial", 12), width=30)
        self.search_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X, padx=5)
        tk.Button(search_frame, text="جستجو", font=("Arial", 12), command=self._search_employee_for_payroll, bg="#03A9F4", fg="white", padx=10, pady=5).pack(side=tk.LEFT, padx=5)

        # Employee Info Display
        self.info_label = tk.Label(self, text="اطلاعات کارمند: ", font=("Arial", 12, "bold"), bg="#f9f9f9", wraplength=700, justify=tk.RIGHT)
        self.info_label.pack(pady=10, padx=20, fill=tk.X)

        # Action Buttons related to other payroll aspects
        other_actions_frame = tk.Frame(self, bg="#f9f9f9")
        other_actions_frame.pack(pady=5, padx=20, fill=tk.X)
        tk.Label(other_actions_frame, text="عملیات پیشرفته حقوق:", font=("Arial", 12, "bold"), bg="#f9f9f9").pack(side=tk.RIGHT, padx=5, pady=5)
        self.manage_aol_button = tk.Button(other_actions_frame, text="مدیریت غیبت، اضافه کار و مرخصی", font=("Arial", 11), command=self._open_absence_overtime_leave_frame, bg="#9C27B0", fg="white", padx=10, pady=5, state="disabled")
        self.manage_aol_button.pack(side=tk.LEFT, padx=5)

        # NEW: Monthly Payroll Calculation Section
        payroll_calc_frame = tk.LabelFrame(self, text="محاسبه و ثبت حقوق ماهانه", bg="#f9f9f9", padx=10, pady=10)
        payroll_calc_frame.pack(pady=10, padx=20, fill=tk.X)

        tk.Label(payroll_calc_frame, text="ماه حقوق (YYYY-MM):", font=("Arial", 10), bg="#f9f9f9").grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.payroll_month_entry = tk.Entry(payroll_calc_frame, font=("Arial", 10), width=15)
        self.payroll_month_entry.insert(0, datetime.now().strftime("%Y-%m"))
        self.payroll_month_entry.grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        tk.Label(payroll_calc_frame, text="مزایا (مثلاً پاداش):", font=("Arial", 10), bg="#f9f9f9").grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.benefits_entry = tk.Entry(payroll_calc_frame, font=("Arial", 10), width=15)
        self.benefits_entry.insert(0, "0")
        self.benefits_entry.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        tk.Label(payroll_calc_frame, text="کسورات (مثلاً جریمه):", font=("Arial", 10), bg="#f9f9f9").grid(row=1, column=3, padx=5, pady=5, sticky="w")
        self.deductions_entry = tk.Entry(payroll_calc_frame, font=("Arial", 10), width=15)
        self.deductions_entry.insert(0, "0")
        self.deductions_entry.grid(row=1, column=2, padx=5, pady=5, sticky="ew")

        tk.Button(payroll_calc_frame, text="محاسبه فیش حقوقی", font=("Arial", 12, "bold"), command=self._calculate_and_record_payroll, bg="#FFC107", fg="black", padx=10, pady=5).grid(row=2, column=0, columnspan=4, pady=10)
        payroll_calc_frame.grid_columnconfigure(0, weight=1)


        # Payroll History Display (for full payslips)
        history_frame = tk.LabelFrame(self, text="سوابق فیش‌های حقوقی", bg="#f9f9f9", padx=10, pady=10)
        history_frame.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)

        self.payroll_history_tree = ttk.Treeview(history_frame, columns=("month", "base_salary", "overtime", "absence", "benefits", "deductions", "loan_deduction", "net_payment", "recorded_date"), show="headings")
        self.payroll_history_tree.pack(fill=tk.BOTH, expand=True)

        self.payroll_history_tree.heading("month", text="ماه", anchor="center")
        self.payroll_history_tree.heading("base_salary", text="حقوق پایه", anchor="center")
        self.payroll_history_tree.heading("overtime", text="اضافه کار (ساعت)", anchor="center")
        self.payroll_history_tree.heading("absence", text="غیبت (ساعت)", anchor="center")
        self.payroll_history_tree.heading("benefits", text="مزایا", anchor="center")
        self.payroll_history_tree.heading("deductions", text="کسورات", anchor="center")
        self.payroll_history_tree.heading("loan_deduction", text="کسر وام", anchor="center")
        self.payroll_history_tree.heading("net_payment", text="خالص پرداخت", anchor="center")
        self.payroll_history_tree.heading("recorded_date", text="تاریخ ثبت فیش", anchor="center")

        self.payroll_history_tree.column("month", width=80, anchor="center")
        self.payroll_history_tree.column("base_salary", width=100, anchor="center")
        self.payroll_history_tree.column("overtime", width=80, anchor="center")
        self.payroll_history_tree.column("absence", width=80, anchor="center")
        self.payroll_history_tree.column("benefits", width=80, anchor="center")
        self.payroll_history_tree.column("deductions", width=80, anchor="center")
        self.payroll_history_tree.column("loan_deduction", width=80, anchor="center")
        self.payroll_history_tree.column("net_payment", width=100, anchor="center")
        self.payroll_history_tree.column("recorded_date", width=120, anchor="center")

        history_yscrollbar = ttk.Scrollbar(history_frame, orient="vertical", command=self.payroll_history_tree.yview)
        history_yscrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.payroll_history_tree.configure(yscrollcommand=history_yscrollbar.set)

        self.payroll_history_tree.bind("<<TreeviewSelect>>", self._on_payslip_select)


        self._clear_payroll_inputs()
        self._load_payroll_history() # Load any existing history on startup

    def _search_employee_for_payroll(self):
        national_id = self.search_entry.get().strip()
        if not national_id:
            messagebox.showwarning("ورودی ناقص", "لطفا کد ملی کارمند را وارد کنید.")
            self._clear_payroll_inputs()
            self.current_employee_id = None
            return

        if not national_id.isdigit() or len(national_id) != 10:
            messagebox.showerror("خطا", "کد ملی باید یک عدد ۱۰ رقمی باشد.")
            self._clear_payroll_inputs()
            self.current_employee_id = None
            return

        emp_data = get_employee_data(national_id)
        if emp_data:
            self.current_employee_id = national_id

            # Re-enable inputs
            self.payroll_month_entry.config(state="normal")
            self.benefits_entry.config(state="normal")
            self.deductions_entry.config(state="normal")
            self.manage_aol_button.config(state="normal") # Enable A/O/L button

            # Display employee info
            self.info_label.config(text=f"کارمند: {emp_data[1]} {emp_data[2]} (کد ملی: {emp_data[0]})\nسمت: {emp_data[3]}\nحقوق ثابت: {emp_data[4]:.0f} تومان", justify=tk.RIGHT)

            self._load_payroll_history() # Load payroll history for the selected employee
        else:
            messagebox.showerror("خطا", "کارمندی با این کد ملی یافت نشد.")
            self._clear_payroll_inputs()
            self.current_employee_id = None

    def _clear_payroll_inputs(self):
        self.search_entry.delete(0, tk.END)
        self.info_label.config(text="اطلاعات کارمند: ")
        self.payroll_month_entry.delete(0, tk.END)
        self.payroll_month_entry.insert(0, datetime.now().strftime("%Y-%m"))
        self.benefits_entry.delete(0, tk.END)
        self.benefits_entry.insert(0, "0")
        self.deductions_entry.delete(0, tk.END)
        self.deductions_entry.insert(0, "0")

        # Disable inputs until employee is selected
        self.payroll_month_entry.config(state="disabled")
        self.benefits_entry.config(state="disabled")
        self.deductions_entry.config(state="disabled")
        self.manage_aol_button.config(state="disabled")

        # Clear payroll history tree
        for item in self.payroll_history_tree.get_children():
            self.payroll_history_tree.delete(item)
        self.payroll_history_tree.insert("", tk.END, values=("", "", "", "", "", "", "", "لطفا ابتدا یک کارمند را جستجو کنید.", ""), tags=('no_data',))
        self.payroll_history_tree.tag_configure('no_data', foreground='gray', font=('Arial', 10, 'italic'))

    def _calculate_and_record_payroll(self):
        if not self.current_employee_id:
            messagebox.showwarning("هشدار", "لطفا ابتدا کارمندی را برای محاسبه حقوق انتخاب کنید.")
            return

        payroll_month = self.payroll_month_entry.get().strip()
        benefits_str = self.benefits_entry.get().strip()
        deductions_str = self.deductions_entry.get().strip()

        if not payroll_month:
            messagebox.showwarning("ورودی ناقص", "لطفا ماه حقوق را وارد کنید.")
            return

        try:
            # Validate YYYY-MM format
            datetime.strptime(payroll_month, "%Y-%m")
        except ValueError:
            messagebox.showerror("خطا", "فرمت ماه حقوق نامعتبر است. لطفا از فرمت YYYY-MM استفاده کنید (مثلاً 2023-01).")
            return

        try:
            benefits = float(benefits_str)
            if benefits < 0:
                messagebox.showerror("خطا", "مزایا نمی‌تواند منفی باشد.")
                return
        except ValueError:
            messagebox.showerror("خطا", "مزایا باید یک عدد معتبر باشد.")
            return

        try:
            deductions = float(deductions_str)
            if deductions < 0:
                messagebox.showerror("خطا", "کسورات نمی‌تواند منفی باشد.")
                return
        except ValueError:
            messagebox.showerror("خطا", "کسورات باید یک عدد معتبر باشد.")
            return

        # Fetch employee's current base salary
        emp_data = get_employee_data(self.current_employee_id)
        if not emp_data:
            messagebox.showerror("خطا", "اطلاعات کارمند یافت نشد.")
            return
        base_salary = emp_data[4]

        # Calculate total overtime and absence for the given month
        total_overtime_hours = get_overtimes_in_month(self.current_employee_id, payroll_month)
        total_absence_hours = get_absences_in_month(self.current_employee_id, payroll_month)

        # --- Loan Deduction Logic ---
        active_loans = get_active_loans(self.current_employee_id)
        total_loan_deduction_for_month = 0
        loan_deductions_detail = []

        if active_loans:
            for loan_id, loan_date, amount, remaining_amount, installment_amount, description in active_loans:
                deduct_this_loan = min(installment_amount, remaining_amount) if installment_amount > 0 else 0
                if deduct_this_loan > 0:
                    total_loan_deduction_for_month += deduct_this_loan
                    loan_deductions_detail.append(f"وام (ID: {loan_id}): {deduct_this_loan:.0f} تومان (باقی‌مانده: {remaining_amount - deduct_this_loan:.0f})")
                    # Update loan remaining amount immediately
                    update_loan_remaining_amount(loan_id, remaining_amount - deduct_this_loan)

        # --- Simple Salary Calculation ---
        HOURLY_BASE_RATE = base_salary / HOURLY_WORK_HOURS_IN_MONTH
        overtime_pay = total_overtime_hours * HOURLY_BASE_RATE * OVERTIME_RATE_FACTOR
        absence_deduction = total_absence_hours * HOURLY_BASE_RATE * ABSENCE_RATE_FACTOR

        # Calculate Net Payment
        net_payment = base_salary + overtime_pay + benefits - absence_deduction - deductions - total_loan_deduction_for_month

        # Prepare payslip details (can be more structured, e.g., JSON)
        payslip_details = (
            f"گزارش فیش حقوقی برای ماه: {payroll_month}\n"
            f"حقوق پایه: {base_salary:.0f} تومان\n"
            f"ساعات اضافه کار: {total_overtime_hours:.1f} ساعت (پاداش: {overtime_pay:.0f} تومان)\n"
            f"ساعات غیبت: {total_absence_hours:.1f} ساعت (کسر: {absence_deduction:.0f} تومان)\n"
            f"مزایا: {benefits:.0f} تومان\n"
            f"کسورات متفرقه: {deductions:.0f} تومان\n"
            f"کسر بابت وام / پیش‌پرداخت: {total_loan_deduction_for_month:.0f} تومان"
        )
        if loan_deductions_detail:
            payslip_details += "\n  - جزئیات کسر وام:\n    " + "\n    ".join(loan_deductions_detail)
        payslip_details += f"\n\nمبلغ خالص پرداخت: {net_payment:.0f} تومان"


        success = record_monthly_payroll_to_db(
            self.current_employee_id, payroll_month, base_salary,
            total_overtime_hours, total_absence_hours, benefits, deductions,
            total_loan_deduction_for_month, net_payment, payslip_details
        )
        if success:
            messagebox.showinfo("موفقیت", "فیش حقوقی با موفقیت ثبت شد.")
            self._load_payroll_history()
            # Refresh employee info to show updated loan status if any
            self._search_employee_for_payroll()
        else:
            messagebox.showerror("خطا", f"فیش حقوقی برای این کارمند در ماه {payroll_month} قبلاً ثبت شده است.")

        # Clear specific entries after calculation, but keep month for consecutive entries
        self.benefits_entry.delete(0, tk.END)
        self.benefits_entry.insert(0, "0")
        self.deductions_entry.delete(0, tk.END)
        self.deductions_entry.insert(0, "0")


    def _load_payroll_history(self):
        for item in self.payroll_history_tree.get_children():
            self.payroll_history_tree.delete(item)

        if self.current_employee_id:
            history = get_payroll_history(self.current_employee_id)
            if not history:
                self.payroll_history_tree.insert("", tk.END, values=("", "", "", "", "", "", "", "هیچ فیش حقوقی یافت نشد.", ""), tags=('no_data',))
                self.payroll_history_tree.tag_configure('no_data', foreground='gray', font=('Arial', 10, 'italic'))
                return

            for record in history:
                self.payroll_history_tree.insert("", tk.END, values=(
                    record[0], # payroll_month
                    f"{record[1]:.0f}", # base_salary_at_time
                    f"{record[2]:.1f}", # overtime_hours
                    f"{record[3]:.1f}", # absence_hours
                    f"{record[4]:.0f}", # benefits
                    f"{record[5]:.0f}", # deductions
                    f"{record[6]:.0f}", # loan_deduction
                    f"{record[7]:.0f}", # net_payment
                    record[9] # recorded_date
                ))
        else:
            self.payroll_history_tree.insert("", tk.END, values=("", "", "", "", "", "", "", "لطفا ابتدا یک کارمند را جستجو کنید.", ""), tags=('no_data',))
            self.payroll_history_tree.tag_configure('no_data', foreground='gray', font=('Arial', 10, 'italic'))

    def _on_payslip_select(self, event):
        selected_item = self.payroll_history_tree.selection()
        if not selected_item:
            return

        item_values = self.payroll_history_tree.item(selected_item[0], 'values')

        payroll_month = item_values[0] # YYYY-MM
        conn = sqlite3.connect(DB_NAME) # Using direct sqlite3 connection here, could add a specific get_payslip_details to database_ops
        cursor = conn.cursor()
        cursor.execute("SELECT payslip_details FROM payroll WHERE employee_national_id = ? AND payroll_month = ?",
                       (self.current_employee_id, payroll_month))
        payslip_details = cursor.fetchone()[0]
        conn.close()

        if payslip_details:
            messagebox.showinfo(f"جزئیات فیش حقوقی {payroll_month}", payslip_details)


    def _open_absence_overtime_leave_frame(self):
        if not self.current_employee_id:
            messagebox.showwarning("هشدار", "لطفا ابتدا کارمندی را برای مدیریت غیبت، اضافه کار و مرخصی جستجو کنید.")
            return
        self.app.show_frame(AbsenceOvertimeLeaveFrame, self.current_employee_id)


class ReportsFrame(tk.Frame):
    def __init__(self, master, app_instance):
        super().__init__(master, bg="#f9f9f9")
        self.app = app_instance

        tk.Label(self, text="گزارش‌گیری", font=("Arial", 16, "bold"), bg="#f9f9f9").pack(pady=20)

        tk.Button(self, text="گزارش لیست کارمندان (اکسل)", font=("Arial", 12, "bold"), command=self._export_employees_to_excel, bg="#008000", fg="white", padx=10, pady=5).pack(pady=10)
        tk.Button(self, text="گزارش خلاصه حقوق و دستمزد (اکسل)", font=("Arial", 12, "bold"), command=self._export_payroll_summary_to_excel, bg="#008000", fg="white", padx=10, pady=5).pack(pady=10)
        tk.Button(self, text="گزارش کامل فیش‌های حقوقی (اکسل)", font=("Arial", 12, "bold"), command=self._export_full_payslips_to_excel, bg="#008000", fg="white", padx=10, pady=5).pack(pady=10)


        tk.Button(self, text="بازگشت به منو", font=("Arial", 12), command=self.app.create_main_menu_frame, bg="#FFC107", fg="black", padx=10, pady=5).pack(pady=30)

    def _export_employees_to_excel(self):
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                               filetypes=[("Excel files", "*.xlsx")],
                                               initialfile="گزارش-کارمندان.xlsx")
        if not filename:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "لیست کارمندان"

            headers = ["کد ملی", "نام", "نام خانوادگی", "سمت", "حقوق ثابت"]
            sheet.append(headers)

            employees_data = get_employee_data()
            for emp in employees_data:
                sheet.append(list(emp)) # Convert tuple to list for append

            workbook.save(filename)
            messagebox.showinfo("موفقیت", f"گزارش لیست کارمندان با موفقیت در '{filename}' ذخیره شد.")
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در ذخیره فایل اکسل: {e}")

    def _export_payroll_summary_to_excel(self):
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                               filetypes=[("Excel files", "*.xlsx")],
                                               initialfile="گزارش-خلاصه-حقوق-دستمزد.xlsx")
        if not filename:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "خلاصه حقوق و دستمزد"

            headers = ["کد ملی", "نام", "نام خانوادگی", "حقوق ثابت", "مجموع پرداختی خالص", "مجموع وام های فعال", "مجموع باقیمانده وام ها"]
            sheet.append(headers)

            employees_data = get_employee_data()
            conn = sqlite3.connect(DB_NAME)
            cursor = conn.cursor()

            for emp in employees_data:
                national_id = emp[0]
                base_salary = emp[4]

                # Sum all net_payment from payroll table for a simplified "Total Paid"
                cursor.execute("SELECT SUM(net_payment) FROM payroll WHERE employee_national_id = ?", (national_id,))
                total_net_paid = cursor.fetchone()[0] or 0

                # Sum active loans for current employee
                cursor.execute("SELECT SUM(amount), SUM(remaining_amount) FROM loans WHERE employee_national_id = ? AND is_active = 1", (national_id,))
                loan_summary = cursor.fetchone()
                total_loan_amount = loan_summary[0] or 0
                total_loan_remaining = loan_summary[1] or 0

                row_data = [
                    national_id,
                    emp[1], # first_name
                    emp[2], # last_name
                    base_salary,
                    total_net_paid,
                    total_loan_amount,
                    total_loan_remaining
                ]
                sheet.append(row_data)

            conn.close()
            # Auto-size columns for better readability
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width


            workbook.save(filename)
            messagebox.showinfo("موفقیت", f"گزارش خلاصه حقوق و دستمزد با موفقیت در '{filename}' ذخیره شد.")
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در ذخیره فایل اکسل: {e}")

    def _export_full_payslips_to_excel(self):
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                               filetypes=[("Excel files", "*.xlsx")],
                                               initialfile="گزارش-فیش‌های-حقوقی-کامل.xlsx")
        if not filename:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "فیش‌های حقوقی کامل"

            headers = ["کد ملی", "نام", "نام خانوادگی", "ماه", "حقوق پایه (زمان فیش)", "ساعت اضافه کار", "ساعت غیبت", "مزایا", "کسورات متفرقه", "کسر وام", "خالص پرداخت", "تاریخ ثبت فیش", "جزئیات فیش"]
            sheet.append(headers)

            conn = sqlite3.connect(DB_NAME)
            cursor = conn.cursor()

            # Fetch all payroll records with employee names
            cursor.execute('''
                SELECT 
                    e.national_id, e.first_name, e.last_name, 
                    p.payroll_month, p.base_salary_at_time, p.overtime_hours, 
                    p.absence_hours, p.benefits, p.deductions, p.loan_deduction, 
                    p.net_payment, p.recorded_date, p.payslip_details
                FROM payroll p
                JOIN employees e ON p.employee_national_id = e.national_id
                ORDER BY e.national_id, p.payroll_month DESC
            ''')
            all_payslips = cursor.fetchall()
            conn.close()

            if not all_payslips:
                messagebox.showwarning("گزارش خالی", "هیچ فیش حقوقی در سیستم ثبت نشده است.")
                return

            for payslip in all_payslips:
                sheet.append(list(payslip))

            # Auto-size columns
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width

            workbook.save(filename)
            messagebox.showinfo("موفقیت", f"گزارش کامل فیش‌های حقوقی با موفقیت در '{filename}' ذخیره شد.")
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در ذخیره فایل اکسل: {e}")